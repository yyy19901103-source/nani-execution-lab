"""Generate 150-row past devicra learning Excel for spec-risk-extractor."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

ROWS = [
    # ── FLANGES / PIPING ──────────────────────────────────────────────────────
    ("All flanges shall be ASME B16.5 Class 300 rating minimum.",
     "[DEVIATION] Customer requires ASME B16.5 Class 300 flanges. Manufacturer proposes JIS B2220 20K flanges as equivalent based on equal pressure-temperature rating. Technical comparison table enclosed. Request written approval.", "D", "Cost", "対象"),
    ("All process nozzle flanges shall be ASME B16.5 Class 600 raised face.",
     "[DEVIATION] Class 600 raised face for all nozzles significantly increases flange weight and bolt length. Manufacturer proposes Class 300 RF for nozzles ≤ 6\" in low-pressure zones (< 35 barg). Request approval of hybrid rating approach.", "D", "Cost / Design", "対象"),
    ("Ring-joint (RTJ) facing is required for all high-pressure connections above 100 barg.",
     "[CLARIFICATION] RTJ facing requires specific groove dimensions per ASME B16.20. Please confirm: (a) ring material (soft iron, 316SS, or Inconel), (b) ring type (R or RX), and (c) whether Manufacturer's standard groove finish (3.2 µm Ra) is acceptable.", "C", "Design", "要確認"),
    ("Piping within the package shall conform to ASME B31.3 Process Piping.",
     "[DEVIATION] Manufacturer's standard for package interconnecting piping is JIS B2301/2302, which meets equivalent safety margins. Full ASME B31.3 requires re-qualification of all in-house weld procedures. Estimated schedule impact: +3 weeks. Request acceptance of JIS piping standard.", "D", "Delivery / Cost", "対象"),
    ("All small-bore connections (≤ 1\") shall be socket-welded, not threaded.",
     "[DEVIATION] Socket-welded connections for all small-bore instrumentation lines increase manhour cost and fabrication time significantly. Manufacturer proposes socket-weld for process lines and NPT threaded for utility/instrument lines ≤ 1\" where pressure ≤ 20 barg and temperature ≤ 150°C. Request approval.", "D", "Cost / Delivery", "対象"),
    ("Minimum pipe wall thickness shall be Schedule 80 for all process lines.",
     "[DEVIATION] Sch.80 for all lines, including low-pressure lube oil and cooling water, is excessive and increases package weight and cost. Manufacturer proposes Sch.40 for services ≤ 10 barg and Sch.80 for > 10 barg. Request approval of tiered schedule selection.", "D", "Cost", "対象"),
    # ── MATERIALS ──────────────────────────────────────────────────────────────
    ("All wetted parts shall comply with NACE MR0175/ISO 15156 for H2S service.",
     "[CLARIFICATION] H2S concentration is listed as 15 ppm in the process data sheet. Please confirm: (a) partial pressure of H2S at design conditions, (b) whether sour service applies at all operating points including startup/shutdown, as this affects material selection and hardness limits.", "C", "Design / Quality", "要確認"),
    ("No cast iron shall be used for any pressure-containing or rotating parts.",
     "[CLARIFICATION] 'No cast iron' is understood to prohibit grey cast iron (GCI). Please confirm: (a) is ductile iron (nodular/spheroidal graphite iron, GJS) also prohibited, (b) are cast steel components acceptable? Clarification needed before materials are ordered.", "C", "Design", "要確認"),
    ("All material certifications shall be EN 10204 Type 3.2.",
     "[DEVIATION] EN 10204 Type 3.2 requires a third-party inspector to co-sign the test report at the mill. For standard commodity items (bolts, gaskets, small fittings), Type 3.1 is the industry norm. Manufacturer proposes Type 3.2 for pressure-retaining and rotating components only, and Type 3.1 for all other items. Request approval.", "D", "Cost / Delivery", "対象"),
    ("Shaft material shall be AISI 17-4PH stainless steel, H1025 condition.",
     "[CLARIFICATION] 17-4PH H1025 is specified. Please confirm: (a) is H1075 acceptable as an alternative with slightly lower hardness but better corrosion resistance in H2S, (b) does customer have a preferred independent test lab for age-hardening verification?", "C", "Design / Quality", "要確認"),
    ("Impellers shall be machined from a single billet; no welded impellers.",
     "[CLARIFICATION] 'No welded impellers' is confirmed for conventionally welded impellers. Please clarify: are electron-beam welded (EBW) impellers acceptable? EBW is routinely used for narrow-passage high-pressure stages with no filler metal, and is a proven aerospace-grade process.", "C", "Design", "要確認"),
    ("Bolting shall be A193 B7 / A194 2H for all pressure-boundary flanges.",
     "[DEVIATION] A193 B7 studs with A194 2H nuts are Manufacturer's standard and are accepted. However, for connections in H2S service, B7 hardness must not exceed HRC 26 per NACE MR0175. Manufacturer will verify all B7 lots against hardness limit and document in MTR. No schedule impact expected.", "D", "Quality", "対象"),
    ("All gaskets shall be spiral-wound type with 316L inner ring and outer guide ring.",
     "[DEVIATION] Spiral-wound gaskets with 316L winding and outer guide ring are accepted as standard. However, for Class 150 low-pressure non-critical flanges, Manufacturer proposes compressed fiber gaskets (non-asbestos) as a cost-saving measure. Class 300 and above to use spiral-wound. Request approval.", "D", "Cost", "対象"),
    ("All external bolting shall be 316 stainless steel, minimum Grade B8.",
     "[DEVIATION] 316SS B8 bolting for high-temperature flanged joints above 120°C is prone to galling (cold-welding) during assembly and disassembly. Manufacturer proposes A193 B7 alloy steel with Molykote anti-seize lubricant for connections above 120°C, and B8M for ambient-temperature connections.", "D", "Design / Quality", "対象"),
    # ── TESTING ────────────────────────────────────────────────────────────────
    ("Performance test per API 610 Table 2 tolerances; any deviation requires re-test.",
     "[DEVIATION] Mandatory re-test after any tolerance exceedance is impractical and creates schedule risk. Manufacturer proposes: deviations within 5% of the tolerance band may be resolved by engineering analysis with Purchaser's written approval, without physical re-test.", "D", "Cost / Delivery", "対象"),
    ("Hydrostatic test at 1.5× MAWP with Purchaser hold point (H).",
     "[CLARIFICATION] Hydrostatic test at 1.5× MAWP is accepted. Please confirm: (a) test medium (water with rust inhibitor), (b) minimum test duration (Manufacturer's standard is 30 minutes after pressurization), (c) whether pressure gauge calibration certificates are required.", "C", "Quality", "要確認"),
    ("NPSH test per API 610 Annex F shall be conducted at rated speed.",
     "[DEVIATION] API 610 Annex F NPSH test is required. Manufacturer's test facility can conduct this test but it requires additional test loop setup time (+5 days). Please confirm this is a contractual requirement (H-point) and not optional, so Manufacturer can plan schedule accordingly.", "D", "Delivery", "対象"),
    ("Strip-down inspection after mechanical run test is mandatory.",
     "[CLARIFICATION] Strip-down inspection scope: please confirm if this includes (a) bearing removal and measurement, (b) seal removal and inspection, (c) wear ring clearance measurement, (d) visual inspection only. Manufacturer's standard strip-down covers (a) through (d). Clarification needed for ITP preparation.", "C", "Quality", "要確認"),
    ("String test (compressor + gearbox + driver assembled) shall be conducted.",
     "[CLARIFICATION] String test with steam turbine driver: please confirm (a) steam supply conditions available at Manufacturer's facility (design: 42 barg / 400°C; available: 35 barg / 360°C max), (b) whether corrected performance by calculation is acceptable if steam conditions are lower than design.", "C", "Cost / Delivery", "要確認"),
    ("Vibration measurement per API 670 during mechanical run test.",
     "[CLARIFICATION] API 670 vibration measurement is standard. Please confirm: (a) alert and trip setpoints at the time of test (are trip setpoints to be active during factory test?), (b) whether spectrum analysis (FFT) data is required in addition to overall levels.", "C", "Quality", "要確認"),
    ("All sub-vendor test results must be submitted to Purchaser for approval before shipment.",
     "[DEVIATION] Requiring Purchaser approval of all sub-vendor test results before shipment creates a bottleneck for standard commodity items (fasteners, standard valves). Manufacturer proposes: approval required for all critical sub-vendors (bearings, seals, instrumentation); Manufacturer's self-certification for standard commodities.", "D", "Delivery", "対象"),
    ("Lube oil system cleanliness to ISO 4406 level 16/14/11 minimum.",
     "[DEVIATION] ISO 4406 level 16/14/11 requires extended flushing time (estimated 3-5 days per system) with multiple oil changes. Manufacturer proposes 17/15/12 per API 614 minimum requirement. If 16/14/11 is mandatory, schedule impact of +14 calendar days must be formally acknowledged.", "D", "Cost / Delivery", "対象"),
    ("Factory Acceptance Test (FAT) of control panel minimum 2 full days; Purchaser to attend.",
     "[CLARIFICATION] 2-day FAT for LCP is accepted. Please advise: (a) number of Purchaser's attendees (for site pass preparation), (b) whether sub-contractor (DCS supplier) attendance is required at FAT, (c) preferred FAT schedule — Manufacturer proposes Week 40.", "C", "Delivery", "要確認"),
    ("Rotor overspeed test at 115% of MCS; results to be witnessed.",
     "[CLARIFICATION] Rotor overspeed test at 115% MCS is standard practice. Please confirm: (a) is this a destructive trip test (rotor reaches 115% under its own momentum) or a regulated over-speed test using governor, (b) is Purchaser witness required on-site, or is video recording acceptable?", "C", "Quality / Cost", "要確認"),
    # ── INSPECTION ─────────────────────────────────────────────────────────────
    ("Third-party inspection by Purchaser's nominated TPI agency with unrestricted access.",
     "[CLARIFICATION] 'Unrestricted access' conflicts with Manufacturer's classified area security policy (defense-related zones). Manufacturer proposes: TPI and Purchaser's inspector may access all pump/compressor manufacturing areas freely; escort required for restricted zones with 24-hour advance notice.", "C", "Quality", "要確認"),
    ("Purchaser reserves the right to inspect at any sub-vendor facility.",
     "[CLARIFICATION] Sub-vendor inspection access depends on each sub-vendor's own facility policies. Manufacturer will facilitate TPI access to Tier-1 critical sub-vendors (bearings, seals, gearbox). Access to Tier-2 commodity suppliers (fasteners, standard fittings) is at sub-vendor's discretion. Is this scope of access acceptable?", "C", "Quality", "要確認"),
    ("100% radiographic examination for all butt welds on pressure-retaining parts.",
     "[DEVIATION] 100% RT for all butt welds significantly increases schedule and cost. Manufacturer proposes: 100% RT for P-numbers 1, 3, 4, and 5 materials; 20% random RT for P-number 1 carbon steel welds ≤ 10 barg per ASME Section IX Category D joints. Request approval of tiered examination plan.", "D", "Cost / Delivery", "対象"),
    ("Dye-penetrant examination (PT) on all machined impeller surfaces after final machining.",
     "[CLARIFICATION] PT on all impeller surfaces after machining is accepted. Please confirm: (a) acceptance criteria — ASME Section V Article 6 or equivalent, (b) whether both liquid-penetrant (wet) and fluorescent methods are acceptable, (c) is PT required before or after final balancing?", "C", "Quality", "要確認"),
    ("Hardness testing on all NACE-critical components, HRC 26 maximum.",
     "[CLARIFICATION] Hardness limit HRC 26 per NACE MR0175 is understood for wetted components in H2S service. Please confirm: (a) does this apply to both base metal and heat-affected zone (HAZ) of welds, (b) is Rockwell C scale acceptable or is Brinell (HBW 248 max) required for thicker sections?", "C", "Quality", "要確認"),
    ("Magnetic particle examination (MT) on all forgings before machining.",
     "[DEVIATION] MT on all forgings is accepted for pressure-boundary components. However, for stainless steel forgings (non-magnetic), MT is not applicable. Manufacturer proposes PT as the alternative for austenitic stainless steel forgings. Request confirmation that this substitution is acceptable.", "D", "Quality", "対象"),
    # ── DOCUMENTATION ──────────────────────────────────────────────────────────
    ("Drawing approval cycle: 10 working days for resubmission.",
     "[DEVIATION] 10 working days is insufficient for complex multi-sheet documents (P&IDs, GA drawings, rotor dynamic reports). Manufacturer proposes: 10 WD for single-sheet drawings and data sheets; 15 WD for documents with more than 5 sheets or requiring multi-discipline review.", "D", "Delivery", "対象"),
    ("All drawings shall use Purchaser's title block format.",
     "[DEVIATION] Adopting Purchaser's title block requires conversion of all Manufacturer's CAD templates. This is a significant engineering effort (+2 weeks setup). Manufacturer proposes to use its own title block with Purchaser's project title, document number, and revision box added. Request approval of hybrid approach.", "D", "Cost / Delivery", "対象"),
    ("All documentation shall be in English language only.",
     "[DEVIATION] Manufacturer's internal manufacturing instructions and in-house QC records are primarily in Japanese. Final deliverable documents (drawings, data sheets, O&M manual, test reports) will be in English. Internal shop floor documents available in Japanese only — these are not part of the MDR.", "D", "Quality", "対象"),
    ("Operating and Maintenance Manual to be submitted 6 weeks before shipment.",
     "[DEVIATION] Completing the O&M manual 6 weeks before shipment is not feasible when final performance test data (needed for O&M appendix) is collected only 2 weeks before shipment. Manufacturer proposes: preliminary O&M manual 6 weeks before shipment; final O&M with test data within 4 weeks after shipment.", "D", "Delivery", "対象"),
    ("Vendor shall provide all drawings in native CAD format (DWG or DXF).",
     "[DEVIATION] Manufacturer's CAD system is CATIA V5 (3D) with 2D output in PDF and CATIA CATDrawing format. DWG/DXF export is available but may result in loss of layer structure and some annotation formatting. Manufacturer proposes PDF as primary deliverable and DXF on request, with no guarantee of format fidelity.", "D", "Quality", "対象"),
    ("Vendor shall update the Master Document Register (MDR) monthly.",
     "[CLARIFICATION] Monthly MDR updates are accepted. Please confirm the required MDR format: (a) Manufacturer's standard Excel-based register, or (b) integration with Purchaser's document management system (EDMS). If EDMS integration is required, Manufacturer will need EDMS login credentials and training — additional cost may apply.", "C", "Cost / Delivery", "要確認"),
    ("Vendor shall provide a Deviation Register at bid stage listing all specification deviations.",
     "[CLARIFICATION] Deviation Register at bid stage is provided (refer to this document). Please confirm: (a) is Purchaser's written approval of each deviation required before manufacture, or is Purchaser's silence after 10 WD considered acceptance, (b) are deviations identified post-order handled by the same register or a separate TQ process?", "C", "Delivery", "要確認"),
    # ── DELIVERY / SCHEDULE ────────────────────────────────────────────────────
    ("Required delivery within 28 weeks from Purchase Order.",
     "[DEVIATION] 28-week delivery for a complete API 610 pump with full performance test is not achievable. Key long-lead items: motor 14 weeks, shaft forging 12 weeks, mechanical seal 12 weeks. Minimum realistic schedule is 34 weeks. Manufacturer requests schedule relaxation or authorization to order long-lead items before full PO.", "D", "Delivery", "対象"),
    ("Liquidated damages: 0.5% per week, maximum 10% of contract value.",
     "[CLARIFICATION] LD clause at 0.5%/week up to 10% is noted. Please confirm: (a) LD calculation base is the total contract value including spare parts, or excluding spare parts, (b) is there a 'force majeure' clause covering natural disasters and supply chain disruptions, (c) LD is the sole remedy for late delivery.", "C", "Cost", "要確認"),
    ("Vendor shall submit a Level 3 schedule in Primavera P6 format.",
     "[DEVIATION] Manufacturer uses Microsoft Project for project scheduling. Primavera P6 is not available. Manufacturer proposes submission of Level 3 schedule as Microsoft Project .mpp file plus PDF export. Monthly updates will include critical path analysis. Request acceptance of this format.", "D", "Delivery", "対象"),
    ("Partial shipments require Purchaser's prior written approval.",
     "[CLARIFICATION] Purchaser approval required for partial shipments is noted. Manufacturer anticipates two shipments: (1) instrumentation/LCP (Week 24), (2) main package (Week 28). Please pre-approve this shipping plan in the PO or notify if a separate approval request is needed for each shipment.", "C", "Delivery", "要確認"),
    ("Minimum 6 weeks' advance notice of readiness for pre-shipment inspection.",
     "[DEVIATION] 6 weeks' advance notice for pre-shipment inspection is challenging when the schedule is compressed near the delivery date. Manufacturer proposes 4 weeks' advance notice as a standard, with best-effort to extend to 6 weeks when schedule allows. Request acceptance of 4-week minimum.", "D", "Delivery", "対象"),
    # ── NOISE / VIBRATION ──────────────────────────────────────────────────────
    ("Noise level shall not exceed 85 dB(A) at 1 m from the package boundary.",
     "[DEVIATION] Guaranteed noise level of 85 dB(A) at 1 m under full-load conditions cannot be met without an acoustic enclosure for this compressor size. Predicted level: 88-91 dB(A). Manufacturer proposes supply with acoustic enclosure at additional cost. Alternatively, if site hearing protection is used, a 90 dB(A) limit per ISO 1680 may be acceptable.", "D", "Cost / Design", "対象"),
    ("Vibration alert setpoint: 25.4 µm pk-pk; trip setpoint: 50.8 µm pk-pk.",
     "[CLARIFICATION] Alert and trip setpoints are noted. Please confirm: (a) are these values to be active during mechanical run test at the factory (if yes, test may be interrupted by alert), (b) are these values fixed or adjustable in the field via the MMS software, (c) are separate setpoints needed for startup versus steady-state operation?", "C", "Design", "要確認"),
    ("Critical speeds shall be at least 15% removed from operating speed range.",
     "[CLARIFICATION] 15% margin from operating speed range is understood. Please confirm: does 'operating speed range' mean minimum governor speed to trip speed, or rated speed ±10%? The margin calculation changes significantly depending on interpretation.", "C", "Design", "要確認"),
    ("Maximum bearing temperature (white metal): 110°C at rated conditions.",
     "[CLARIFICATION] 110°C limit for babbitt journal bearings is accepted. Please confirm: (a) is this the oil film temperature or the bearing shell temperature, (b) is 110°C the alarm or the trip setpoint, (c) what is the trip setpoint (Manufacturer's standard: alarm 110°C / trip 120°C)?", "C", "Design", "要確認"),
    # ── LUBE OIL SYSTEM ─────────────────────────────────────────────────────────
    ("Lube oil reservoir internal coating: two-component epoxy, minimum 200 µm DFT.",
     "[DEVIATION] Internal epoxy coating of lube oil reservoir risks delamination of coating particles entering the oil circuit and damaging bearings. Manufacturer's standard: uncoated pickled carbon steel with nitrogen blanket for long-term storage. Alternatively, 316SS internal shell welded to CS outer shell. Request approval of alternative.", "D", "Design / Quality", "対象"),
    ("Oil coolers shall be shell-and-tube type with cooling water on tube side.",
     "[DEVIATION] Shell-and-tube coolers with cooling water on tube side is accepted as standard. However, for high-fouling cooling water (high TDS or biological growth), Manufacturer proposes plate-and-frame heat exchangers as an alternative — easier to clean, smaller footprint. Please advise preference or site cooling water quality data.", "D", "Design", "対象"),
    ("Dual lube oil filters, changeable under pressure, 10 µm absolute.",
     "[CLARIFICATION] 10 µm absolute filtration is specified. Please confirm: (a) is 10 µm absolute or 10 µm nominal (significant difference in filter efficiency), (b) are disposable filter elements or cleanable elements preferred, (c) is a differential pressure indicator sufficient or is a DP transmitter with alarm required?", "C", "Design", "要確認"),
    ("Lube oil supply temperature to bearings: 45°C ±3°C.",
     "[CLARIFICATION] 45°C ±3°C supply temperature is accepted. Please confirm: (a) this is the supply temperature at the bearing housing inlet, not at the cooler outlet, (b) is a temperature control valve (TCV) on the oil cooler bypass required, (c) what is the minimum oil temperature before startup is permitted?", "C", "Design", "要確認"),
    ("Oil reservoir minimum hold-up volume: 5 minutes at rated flow.",
     "[CLARIFICATION] 5-minute hold-up is the minimum per API 614. Manufacturer's standard is 7-8 minutes to allow adequate time for coasting-down after shutdown. Please confirm if the 5-minute minimum is a hard limit or if a larger reservoir is acceptable (note: larger reservoir increases package footprint and weight).", "C", "Design / Cost", "要確認"),
    # ── DRY GAS SEALS ──────────────────────────────────────────────────────────
    ("Dry gas seals: tandem-type per API 692.",
     "[CLARIFICATION] Tandem seal arrangement is accepted. Please confirm the secondary containment arrangement: (a) vented to atmosphere with monitoring (Plan 72), (b) vented to flare (Plan 75), or (c) buffered with inert gas (Plan 74)? The P&ID cannot be completed without this decision.", "C", "Design", "要確認"),
    ("Primary seal gas flow: maximum 50 Nm³/h per seal.",
     "[DEVIATION] Based on thermodynamic calculation, primary seal gas consumption at rated conditions is estimated at 55-60 Nm³/h per seal for this casing geometry. Manufacturer proposes 65 Nm³/h as the contract guarantee value. Purchaser should confirm that utility supply capacity accommodates this revised value.", "D", "Design", "対象"),
    ("Seal gas supply pressure: minimum 0.5 bar above process pressure at the seal.",
     "[CLARIFICATION] The 0.5 bar delta-P specification: please confirm this is required at the seal itself (not at the supply header). Supply line pressure drop must be included in the utility header sizing. Also confirm minimum and maximum utility supply pressure range so Manufacturer can size the pressure regulator.", "C", "Design", "要確認"),
    # ── STEAM TURBINE ──────────────────────────────────────────────────────────
    ("Steam turbine trip functions shall be hardwired, not via DCS.",
     "[CLARIFICATION] Hardwired trip functions are accepted as standard. Please confirm the complete trip list to be hardwired: Manufacturer's standard includes overspeed, low lube oil pressure, high vibration, high axial displacement, and high exhaust temperature. Are any additional trips required by Purchaser's safety concept?", "C", "Design", "要確認"),
    ("Condensing steam turbine; exhaust to surface condenser.",
     "[CLARIFICATION] Condensing turbine exhaust: please confirm (a) condenser design pressure (Manufacturer's calculation is based on 0.1 bara — please confirm), (b) whether the condenser is in Manufacturer's scope or Purchaser-supplied, (c) hotwell pump and ejector system scope boundary.", "C", "Design", "要確認"),
    ("Turning gear for cool-down rotation after shutdown.",
     "[CLARIFICATION] Turning gear is listed as required. Please confirm (a) turning gear is in Vendor scope (not listed in Section 1.1), (b) required turning speed (typically 1-5 RPM for steam turbines of this size), (c) motor-driven or air-motor-driven preferred.", "C", "Cost / Delivery", "要確認"),
    ("Governor shall allow remote speed setpoint via 4-20 mA signal from DCS.",
     "[CLARIFICATION] Remote speed setpoint via 4-20 mA is accepted. Please confirm: (a) signal range mapping (4 mA = minimum speed, 20 mA = maximum speed, or reverse), (b) fail-safe action on 4-20 mA signal loss (hold last speed, ramp to minimum, or trip), (c) is speed ramp rate adjustable from DCS?", "C", "Design", "要確認"),
    # ── PAINTING / PRESERVATION ───────────────────────────────────────────────
    ("External surfaces: epoxy zinc-rich primer + epoxy MIO intermediate + PU topcoat, DFT ≥ 250 µm.",
     "[DEVIATION] Manufacturer's standard paint system (zinc-rich primer + epoxy intermediate + PU topcoat, 240 µm) is marginally below the 250 µm requirement. Manufacturer proposes increasing intermediate coat by 10 µm to meet the minimum. No cost impact; minor schedule adjustment required for additional coat flash-off time.", "D", "Quality", "対象"),
    ("Final color: RAL 5015 (sky blue) for all package equipment.",
     "[CLARIFICATION] RAL 5015 is accepted for the main package. Please confirm: (a) is the lube oil console the same color as the main package, (b) what color is required for the local control panel (LCP) — typically RAL 7035 light grey for electrical panels, (c) are color samples to be submitted for approval?", "C", "Quality", "要確認"),
    ("All machined surfaces shall be protected with rust-inhibitor; no plastic flange caps.",
     "[DEVIATION] Manufacturer uses plastic flange protectors for shipping as they are lightweight and provide adequate protection for transit periods up to 6 months. For longer storage, steel blanking flanges are used. Please clarify the expected storage period at site — if less than 6 months, Manufacturer requests acceptance of plastic caps.", "D", "Quality", "対象"),
    ("Packing shall withstand 3 months outdoor storage.",
     "[CLARIFICATION] 3 months outdoor storage is noted. Please confirm the expected storage conditions at the destination site: (a) tropical climate (high humidity, high temperature), (b) coastal location (salt-laden atmosphere), (c) whether packages will be stored under cover (roof) or fully exposed to rain and sun.", "C", "Quality / Cost", "要確認"),
    ("N2 blanket (positive pressure 0.2 barg) for all gas-path internals during ocean freight.",
     "[DEVIATION] Maintaining 0.2 barg N2 positive pressure for the full ocean freight duration (estimated 6 weeks) requires a self-contained N2 supply system or on-vessel top-up. Manufacturer proposes N2 purge and seal with relief device set at 0.1 barg, which is the standard for long-distance shipping. Request approval.", "D", "Quality", "対象"),
    # ── HSE ────────────────────────────────────────────────────────────────────
    ("Lifting lugs proof-tested to 150% of maximum lift weight, certified by independent body.",
     "[DEVIATION] Physical proof testing at 150% of assembled package weight (estimated 38 tons, so test load: 57 tons) creates significant risk of structural damage to the assembled package. Manufacturer proposes FEA-based certification per JIS B8817 (125% equivalent analysis) submitted to independent structural certification body. Request acceptance.", "D", "Design / Quality", "対象"),
    ("No asbestos-containing materials; provide declaration of compliance.",
     "[CLARIFICATION] Manufacturer confirms no asbestos in currently procured materials. However, some legacy gasket suppliers in Japan historically used chrysotile gaskets. Please confirm: is a general declaration from Manufacturer sufficient, or are individual component-level MSDS/SDS certificates required for all gaskets and packing materials?", "C", "Quality", "要確認"),
    ("ISO 45001 and ISO 14001 certifications to be submitted at bid.",
     "[DEVIATION] Manufacturer holds ISO 14001 and ISO 9001. ISO 45001 certification audit is scheduled Q3 2026 (currently in ISO 45001 transition from OHSAS 18001). Manufacturer submits OHSAS 18001 certificate plus ISO 45001 gap analysis and implementation schedule. Request acceptance of interim arrangement.", "D", "Quality", "対象"),
    ("All grease fittings shall be compatible with bio-degradable grease per ISO 15380 Type L-EAL.",
     "[CLARIFICATION] Bio-degradable grease compatibility: please confirm whether bio-degradable grease is required for the entire package or only for bearings in environmentally sensitive areas. Standard Zerk fittings (1/4\"-28 UNF) accept all grease types including EAL grades. Manufacturer's standard grease (Mobilux EP2) is not EAL-rated.", "C", "Quality / Cost", "要確認"),
    # ── INSTRUMENTATION ────────────────────────────────────────────────────────
    ("All transmitters shall be HART-compatible, 4-20 mA output.",
     "[CLARIFICATION] HART compatibility is accepted as standard. Please confirm: (a) HART version required (HART 5, 6, or 7 — Manufacturer's standard is HART 7), (b) whether a HART handheld communicator is required in the spare parts list, (c) are HART pass-through capability requirements for DCS integration confirmed?", "C", "Design", "要確認"),
    ("Three speed transmitters per shaft; 2-out-of-3 voting logic for overspeed trip.",
     "[CLARIFICATION] 2oo3 voting logic for overspeed trip is accepted as best practice. Please confirm: (a) is this voting to be implemented in the dedicated ESD system, the MMS, or the PLC in the LCP, (b) is a manual test function (simulate one-out-of-three failure) required in the HMI?", "C", "Design", "要確認"),
    ("Process instrument flanges: ASME B16.5 Class 300 RF; no threaded connections for process measurements.",
     "[DEVIATION] Class 300 flanged connections for all process instruments significantly increase weight and cost on small-bore connections. Manufacturer proposes: Class 300 RF for all process connections ≥ 2\" and in H2S or toxic service; 3/4\" NPT for utility services (lube oil pressure, cooling water) ≤ 10 barg. Request approval.", "D", "Cost", "対象"),
    ("Anti-surge control via dedicated anti-surge controller (ASC), not DCS.",
     "[CLARIFICATION] Dedicated ASC (separate from DCS) is accepted. Please advise: (a) preferred make/model of ASC (Compressor Controls Corporation, Woodward, or other), (b) whether the ASC should be housed in the LCP or a separate panel, (c) ASC communication protocol to DCS (Modbus RTU, Profibus DP, or hardwired analog only).", "C", "Design / Cost", "要確認"),
    ("Machine monitoring system per API 670, stand-alone (not integrated with DCS).",
     "[CLARIFICATION] Stand-alone MMS per API 670 is accepted. Please confirm: (a) preferred MMS supplier (Bently Nevada, Emerson, or equivalent), (b) whether Modbus or OPC-UA communication to DCS is required for trending data, (c) is a dedicated UPS power supply for the MMS required?", "C", "Design / Cost", "要確認"),
    # ── GEARBOX ───────────────────────────────────────────────────────────────
    ("Gearbox service factor: minimum 1.25 on rated power.",
     "[CLARIFICATION] Service factor 1.25 on rated power is accepted. Please confirm: (a) is the service factor applied to the driver rated power or the compressor absorbed power, (b) does the service factor also apply to the gear tooth contact stress (Hertz stress) limit in API 613 rating calculations?", "C", "Design", "要確認"),
    ("Gearbox noise level shall not exceed 85 dB(A) at 1 m under full-load.",
     "[DEVIATION] Gearbox noise at 85 dB(A) at 1 m is extremely low for integrally geared units at this power level. Typical noise for this type is 88-92 dB(A). To achieve 85 dB(A), full acoustic enclosure is required. Manufacturer proposes 90 dB(A) as a more realistic guarantee, with site hearing protection as the engineering control.", "D", "Cost / Design", "対象"),
    ("Gear tooth contact pattern to be checked and documented during assembly.",
     "[CLARIFICATION] Contact pattern check is standard practice. Please confirm: (a) is blue-print contact pattern inspection or dynamic (running) contact pattern check required, (b) what is the minimum acceptable contact pattern (Manufacturer's standard: ≥ 70% face width, ≥ 50% tooth depth), (c) is a photograph of the contact pattern required in the QC dossier?", "C", "Quality", "要確認"),
    # ── HEAT EXCHANGERS ───────────────────────────────────────────────────────
    ("Gas coolers: ASME Section VIII Div.1, U-stamp; process gas on shell side.",
     "[CLARIFICATION] ASME U-stamp with process gas on shell side is accepted. Please confirm: (a) TEMA class (R, C, or B) required, (b) is an expansion joint required on the shell for thermal growth, (c) are removable tube bundles required for cleaning access?", "C", "Design / Quality", "要確認"),
    ("Cooler tube material: ASTM A249 TP316L.",
     "[DEVIATION] TP316L tubes are specified. For cooling water with high chloride content (> 200 ppm Cl-), TP316L is susceptible to stress corrosion cracking. Manufacturer recommends ASTM A249 TP317L or duplex stainless steel (UNS S31803) if cooling water chloride data cannot be confirmed below 100 ppm. Request discussion.", "D", "Design / Quality", "対象"),
    ("Fouling factor: TEMA standard + 25% additional margin on dirty U-value.",
     "[DEVIATION] Additional 25% margin on dirty U-value leads to over-sized coolers. Risk: gas outlet temperature may drop below dew point, causing condensation and liquid carry-over into the compressor. Manufacturer proposes standard TEMA fouling factors with no additional margin, and requests process engineer review of minimum gas temperature.", "D", "Design", "対象"),
    ("CIP (Clean-in-Place) nozzles: N2 purge inlet and drain on each cooler.",
     "[CLARIFICATION] CIP nozzles (N2 purge inlet + drain) are included as standard. Please confirm: (a) N2 purge connection size and rating, (b) drain nozzle size and whether drain connects to Purchaser's closed drain header, (c) whether a vent nozzle is also required for CIP procedure.", "C", "Design", "要確認"),
    # ── SPARE PARTS ────────────────────────────────────────────────────────────
    ("Commissioning spares for 24 months of continuous operation in base scope.",
     "[CLARIFICATION] 24-month commissioning spares included in base scope is noted. Please confirm the definition: (a) spares consumed during first startup and commissioning only (e.g., gaskets, O-rings, filter elements replaced during initial flush), or (b) a full set of running spares sufficient for 24 months of unplanned maintenance?", "C", "Cost", "要確認"),
    ("Installed spare rotor, fully assembled with seals, in transport frame with N2 blanket.",
     "[CLARIFICATION] Installed spare rotor scope: please confirm (a) spare rotor includes its own set of bearings (journal and thrust), (b) spare rotor balance grade (same as operating rotor, ISO G1.0), (c) transport frame material (carbon steel or aluminum), (d) whether a rotor storage/handling procedure document is required.", "C", "Cost / Quality", "要確認"),
    ("2-year and 5-year recommended spare parts list with unit prices at bid stage.",
     "[CLARIFICATION] Recommended spare parts list at bid stage is provided. Please confirm: (a) this is non-binding (for planning purposes only), (b) are prices firm for 12 months from bid submission, (c) is an electronic spare parts catalog (HTML or PDF with exploded view drawings) also required?", "C", "Cost", "要確認"),
    ("Special tools and commissioning tools included in scope.",
     "[CLARIFICATION] Special tools scope: please confirm what is included. Manufacturer's standard list: coupling alignment fixture, bearing puller, seal installation tools, coupling spacer removal tool. Are any process-specific tools required (e.g., gas seal installation jig, impeller torque wrench)?", "C", "Cost", "要確認"),
    # ── ELECTRICAL ─────────────────────────────────────────────────────────────
    ("All electrical equipment: Class I Division 2 Group D (NEC); IEC Zone 2 requires approval.",
     "[DEVIATION] Manufacturer's standard electrical components are certified to IEC 60079 (ATEX/IECEx Zone 2) as the primary standard. NEC Class I Div. 2 equivalent certification (ETL, UL) adds cost and lead time (+4 weeks for some components). Request acceptance of IEC 60079 Zone 2 as equivalent, subject to Purchaser's EX classification authority confirmation.", "D", "Cost / Delivery", "対象"),
    ("All signal cables in metallic conduit; flexible conduit limited to final 500 mm.",
     "[DEVIATION] Manufacturer's standard: rigid metallic conduit for all fixed runs; flexible metal conduit (FMC) for final connection to instruments (up to 750 mm). 500 mm limit on flexible conduit is shorter than Manufacturer's standard practice. Manufacturer requests acceptance of 750 mm flexible section at instrument connections.", "D", "Design", "対象"),
    ("Motor supply cable entry from below (bottom cable entry) to local junction box.",
     "[CLARIFICATION] Bottom cable entry for the LJB is preferred for cable segregation. Please confirm: (a) conduit entry size (depends on cable OD, to be confirmed at detailed design), (b) number of cables entering the LJB (estimated 12-16 multi-pair cables), (c) is stainless steel conduit required or is galvanized steel acceptable?", "C", "Design", "要確認"),
    # ── WEIGHT / DIMENSIONS ───────────────────────────────────────────────────
    ("Maximum assembled package weight: 30 metric tons.",
     "[DEVIATION] Estimated assembled package weight: 38 metric tons, exceeding the 30-ton limit. Manufacturer proposes factory-tested sub-assembly shipment: compressor + gearbox + turbine on base frame (max 26 tons), lube oil console separately (max 8 tons). Site bolt-up alignment procedure will be provided. Request approval.", "D", "Cost / Delivery", "対象"),
    ("Maximum package footprint: 8,000 mm × 3,500 mm.",
     "[DEVIATION] Preliminary GA indicates package footprint: 9,200 mm × 3,300 mm. Width (3,300 mm) is within limit; length (9,200 mm) exceeds 8,000 mm by 1,200 mm. Manufacturer is optimizing cooler arrangement to reduce length. Final GA to be submitted for approval within 4 weeks of order.", "D", "Design", "対象"),
    ("Package height shall not exceed 4,500 mm including all instruments and connections.",
     "[CLARIFICATION] 4,500 mm height limit noted. Preliminary estimate: 4,200 mm to top of lube oil reservoir vent. Please confirm: (a) does height include the acoustic enclosure if supplied, (b) are flexible hoses and instrument tubing within the 4,500 mm envelope or excluded?", "C", "Design", "要確認"),
    # ── VENDOR DATA / COORDINATION ────────────────────────────────────────────
    ("Vendor shall confirm motor shaft dimensions within 14 days of order.",
     "[DEVIATION] Motor shaft dimensions are set by the motor manufacturer at design freeze, typically 8-10 weeks after order. Manufacturer can provide estimated shaft dimensions (based on IEC/NEMA standard for this kW rating) within 14 days, but final dimensions confirmed at 10 weeks. Motor procurement should note this timing.", "D", "Delivery", "対象"),
    ("Vendor shall coordinate with DCS supplier for all communication interface points.",
     "[CLARIFICATION] DCS supplier coordination is accepted. Please provide: (a) DCS supplier name and project engineer contact, (b) DCS I/O list template or required signal format, (c) communication protocol (Modbus RTU, Modbus TCP, or OPC-DA/UA). Manufacturer will schedule interface meeting within 3 weeks of order.", "C", "Design", "要確認"),
    ("Vendor shall attend a pre-order kickoff meeting at Purchaser's office.",
     "[CLARIFICATION] Kickoff meeting at Purchaser's office is accepted. Please confirm (a) preferred date (Manufacturer proposes within 3 weeks of PO), (b) meeting agenda scope (Manufacturer will prepare action item register), (c) are sub-vendors (seal supplier, gearbox supplier) expected to attend?", "C", "Delivery", "要確認"),
    ("Vendor shall provide HAZOP action item responses within 21 days of HAZOP study.",
     "[CLARIFICATION] HAZOP response within 21 days is noted. Please confirm: (a) will Manufacturer be invited to attend the HAZOP (preferred for real-time technical input), (b) is the HAZOP action item register formatted in a specific template, (c) are HAZOP responses subject to Purchaser approval before implementation?", "C", "Quality / Design", "要確認"),
    # ── STANDARDS COMPLIANCE ──────────────────────────────────────────────────
    ("API 617 Chapter 2 (integrally geared compressors) shall govern over Chapter 1 where conflicts exist.",
     "[CLARIFICATION] Chapter 2 precedence over Chapter 1 is noted. Please confirm: (a) are there specific clauses in Chapter 1 that Purchaser wishes to apply even where Chapter 2 conflicts (e.g., rotor dynamic analysis requirements), (b) is Level 2 rotor dynamic analysis per API 617 Annex I mandatory or recommended?", "C", "Design", "要確認"),
    ("ASME Sec. VIII Div. 1 with U-stamp for all pressure vessels; no PED CE mark required.",
     "[DEVIATION] Manufacturer's pressure vessels are certified to PED (Pressure Equipment Directive) CE marking as primary compliance path, with ASME U-stamp available as an option at additional cost (+4 weeks, +3% on vessel price per unit). Request Purchaser to confirm whether ASME U-stamp is mandatory or if PED CE mark is acceptable.", "D", "Cost / Delivery", "対象"),
    ("TEMA Class R for all shell-and-tube heat exchangers.",
     "[DEVIATION] TEMA Class R (most stringent, heavy process industry) is accepted for gas coolers. However, Manufacturer proposes TEMA Class C (commercial) for lube oil coolers and seal gas heaters, which are low-pressure utility services. TEMA R for utility coolers adds cost without commensurate benefit.", "D", "Cost", "対象"),
    ("All pressure relief valves (PRVs) to be API 520/526 certified.",
     "[CLARIFICATION] API 520/526 PRVs are standard. Please confirm: (a) PRVs are in Manufacturer's scope (not listed in Section 1.1), (b) set pressure and orifice size to be determined by Manufacturer based on MAWP, (c) is a thermal relief valve required on the lube oil circuit if the system can be isolated when hot?", "C", "Design / Cost", "要確認"),
    # ── QUALITY MANAGEMENT ────────────────────────────────────────────────────
    ("Vendor's quality system shall be ISO 9001:2015 certified.",
     "[CLARIFICATION] Manufacturer holds ISO 9001:2015 certification. Certificate will be submitted with bid. Please confirm: (a) are sub-vendor quality system requirements to be passed down by Manufacturer to all critical sub-vendors, (b) does Purchaser require a formal quality plan (QCP) specific to this project?", "C", "Quality", "要確認"),
    ("Non-conformance reports (NCRs) shall be issued within 24 hours of discovery and submitted to Purchaser.",
     "[DEVIATION] 24-hour NCR submission is aggressive for NCRs requiring engineering analysis before issuance. Manufacturer proposes: NCR to be opened in internal system within 24 hours (timestamped), with notification to Purchaser; formal NCR document with root cause and disposition submitted within 5 working days.", "D", "Quality", "対象"),
    ("Weld repair requires Purchaser's written approval before execution.",
     "[DEVIATION] Requiring written approval before every weld repair will cause significant schedule delays. Manufacturer proposes: minor cosmetic repairs (surface porosity, undercut < 1 mm) can proceed per Manufacturer's approved repair procedure without prior approval; major repairs (> 25 mm weld length, through-wall defects) require Purchaser's advance approval.", "D", "Delivery / Quality", "対象"),
    ("All welders shall be qualified per ASME Section IX.",
     "[CLARIFICATION] ASME Section IX welder qualification is accepted. Please confirm: (a) does this requirement apply to sub-vendor welding also, (b) are welder qualification records (WQRs) to be submitted for review or available for audit only, (c) is Purchaser's TPI to witness welder qualification tests or review records only?", "C", "Quality", "要確認"),
    # ── MISCELLANEOUS ─────────────────────────────────────────────────────────
    ("Vendor shall provide a certified 3D model (STEP format) of the complete package.",
     "[DEVIATION] Manufacturer's CAD system is CATIA V5. STEP (AP203/AP214) export is available but 3D model quality (surface accuracy, assembly constraints) cannot be guaranteed for all complex assemblies. Manufacturer proposes STEP export as best-effort deliverable, with no warranty of model accuracy for structural or clash analysis.", "D", "Quality", "対象"),
    ("Package shall be designed for minimum 25-year service life.",
     "[CLARIFICATION] 25-year design life is noted as a design intent. Please confirm: (a) is this to be certified by calculation (FEA fatigue analysis), or is adherence to API standards considered sufficient, (b) are there specific cyclic loading conditions (e.g., number of start-stop cycles per year) that drive the fatigue design?", "C", "Design", "要確認"),
    ("Vendor shall provide a Site Acceptance Test (SAT) procedure 4 weeks before commissioning.",
     "[CLARIFICATION] SAT procedure submission 4 weeks before commissioning is accepted. Please confirm: (a) Manufacturer's commissioning engineer attendance at SAT — is this in scope (not included in base price), (b) SAT acceptance criteria — are these the same as FAT criteria or site-specific performance requirements apply?", "C", "Cost / Quality", "要確認"),
    ("Vendor shall warrant the equipment for 24 months from commissioning or 30 months from shipment.",
     "[CLARIFICATION] Warranty period (24 months from commissioning or 30 months from shipment, whichever is earlier) is noted. Please confirm: (a) warranty claims process — lead time for Manufacturer's field engineer to respond on-site, (b) exclusions: consumables (seals, bearings, filter elements) used in normal service, (c) warranty void conditions.", "C", "Cost", "要確認"),
    ("Vendor shall include a training program for Purchaser's operators and maintenance staff.",
     "[CLARIFICATION] Operator/maintenance training program: please confirm (a) training to be conducted at Manufacturer's facility (Japan) or at site, (b) number of trainees (Manufacturer's standard: up to 8 persons), (c) training language (English or Japanese with interpreter), (d) is training cost included in base price?", "C", "Cost / Delivery", "要確認"),
    ("Vendor shall provide a preventive maintenance schedule and work instructions.",
     "[CLARIFICATION] Preventive maintenance schedule and instructions are included in the O&M Manual. Please confirm: (a) is a dedicated computerized maintenance management system (CMMS) upload format required (e.g., SAP PM templates), (b) are maintenance procedures to be in step-by-step format with photo illustrations?", "C", "Quality", "要確認"),
    ("Emergency shutdown (ESD) system shall be a separate, dedicated SIL-rated system.",
     "[CLARIFICATION] Dedicated SIL-rated ESD system: please provide the SIL requirement (SIL 1, 2, or 3) from the Safety Requirement Specification (SRS). Manufacturer's LCP PLC can be configured for SIL 1 (IEC 61511). SIL 2 requires a separate certified safety PLC (e.g., Siemens S7 F-series or equivalent) at additional cost.", "C", "Design / Cost", "要確認"),
    ("All pressure gauges shall be glycerin-filled, 100 mm dial, 1% accuracy.",
     "[DEVIATION] Glycerin-filled gauges are Manufacturer's standard. 100 mm dial is accepted. However, 1% full-scale accuracy for glycerin gauges in vibration service is difficult to guarantee continuously — gauges drift. Manufacturer proposes 1.6% accuracy per EN 837-1 Class 1.6 as standard, with 1% accuracy for critical process gauges only.", "D", "Cost / Quality", "対象"),
    ("All pressure transmitters shall be remote-seal type for viscous or corrosive services.",
     "[CLARIFICATION] Remote seal transmitters for viscous/corrosive services are accepted. Please confirm which process connections require remote seals: Manufacturer identifies the following candidate services: (a) crude oil suction (viscous), (b) seal gas differential pressure (small delta-P), (c) lube oil pressure (Manufacturer's view: direct-connected is acceptable). Please advise.", "C", "Design", "要確認"),
    ("Flow measurement: Coriolis meters for all oil and chemical injection services.",
     "[DEVIATION] Coriolis meters are high-accuracy but expensive and require long straight runs. Manufacturer proposes: Coriolis for seal gas and injection chemical services; turbine flowmeter for lube oil return monitoring. Coriolis for lube oil is technically feasible but increases cost by approximately USD 8,000 per point.", "D", "Cost / Design", "対象"),
    ("Control valves shall be fail-safe; confirm fail-open or fail-closed per P&ID.",
     "[CLARIFICATION] Fail-safe position (FO/FC/FL) must be confirmed for each control valve on the P&ID. Manufacturer has assigned preliminary fail-safe positions based on process safety (anti-surge valve: fail-open; isolation valves: fail-closed). Please review and approve the fail-safe position list before valve procurement.", "C", "Design", "要確認"),
    ("Solenoid valves: 24 VDC coil, energize-to-trip (de-energize to fail-safe).",
     "[CLARIFICATION] De-energize-to-fail-safe (de-energize to open for vent valves) is the standard. Please confirm: (a) 24 VDC supply from Manufacturer's UPS or Purchaser's UPS, (b) are explosion-proof solenoids (ATEX/IECEx) required for all locations or only Zone 1/Zone 2 classified areas?", "C", "Design", "要確認"),
    ("Package shall be designed for crane-lift assembly; all lifting trunnions on base frame.",
     "[CLARIFICATION] Crane-lift assembly with lifting trunnions on the base frame is accepted. Please confirm: (a) maximum crane capacity available at the installation site (constrains sub-assembly lift weight), (b) lifting trunnion design basis (dynamic load factor during rigging), (c) are below-hook lifting spreader beams in Manufacturer's scope?", "C", "Design / Cost", "要確認"),
    ("Thermal insulation: all surfaces above 60°C and below 10°C shall be insulated.",
     "[CLARIFICATION] Insulation boundary (> 60°C and < 10°C) is noted. Please confirm: (a) insulation material preference (mineral wool, calcium silicate, or foam glass for below ambient), (b) cladding material (aluminum or 316SS), (c) does insulation extend to instrument impulse lines or process connections only?", "C", "Design / Cost", "要確認"),
    ("Vendor shall use only approved sub-vendors from Purchaser's Approved Vendor List (AVL).",
     "[CLARIFICATION] Purchaser's AVL is referenced. Please provide the AVL for: (a) bearings (Manufacturer's standard: SKF or NSK), (b) mechanical seals (John Crane or Flowserve), (c) instrumentation (Rosemount/Yokogawa), (d) coupling (Rexnord or Lovejoy). Manufacturer will identify any conflicts with current approved suppliers.", "C", "Quality / Delivery", "要確認"),
    ("Vendor shall not subcontract any work without Purchaser's prior written approval.",
     "[DEVIATION] Requiring written approval for all subcontracting would delay procurement of standard commodity items (off-the-shelf valves, standard flanges, pipe fittings). Manufacturer proposes: approval required for major sub-contracts (bearings, seals, instrumentation, gearbox, turbine); Manufacturer's discretion for standard commodity items below USD 5,000.", "D", "Delivery", "対象"),
    ("Fireproofing of all structural steel supports within the package.",
     "[CLARIFICATION] Fireproofing scope: please confirm (a) fire scenario basis for fireproofing design (hydrocarbon pool fire per UL 1709 or cellulosic fire per UL 263), (b) required fire resistance duration (30, 60, or 120 minutes), (c) is intumescent paint acceptable or is concrete/vermiculite cladding required?", "C", "Design / Cost", "要確認"),
    ("Vendor shall provide detailed weight and center-of-gravity data for each shipping piece.",
     "[CLARIFICATION] Weight and COG data for each shipping piece is provided in the shipping drawing. Please confirm format: (a) is a separate weight control report required in addition to shipping drawing, (b) is the COG required in X, Y, Z coordinates relative to a fixed datum, (c) what is the required COG accuracy tolerance (Manufacturer's standard: ±5%)?", "C", "Quality", "要確認"),
    ("All elastomers shall be confirmed suitable for the operating temperature range -10°C to +150°C.",
     "[DEVIATION] Standard Nitrile (NBR) O-rings are rated to -30°C to +120°C, which covers most of the temperature range. However, at +150°C (maximum), FKM (Viton) is required. Manufacturer proposes FKM for all high-temperature locations (> 100°C) and NBR for ambient/cooling water services. Request approval of this material split.", "D", "Design / Quality", "対象"),
    ("All gasketed connections shall use full-face gaskets to prevent flange rotation.",
     "[DEVIATION] Full-face gaskets are required for flat-face flanges (Class 125 cast iron). For raised-face flanges (Class 150 and above), ring-type gaskets are the engineering norm — full-face gaskets on RF flanges can concentrate load unevenly and reduce effective bolt load. Manufacturer proposes ring gaskets for all RF flanges.", "D", "Design", "対象"),
    ("Vendor shall perform a seismic analysis if the installation is in Zone 2B or above.",
     "[CLARIFICATION] Seismic zone is stated as Zone 2B per UBC 1997. Please provide: (a) the design spectral acceleration values (Sa and Sd) for the site, (b) whether a static analysis (equivalent lateral force) or dynamic analysis (response spectrum) is required, (c) the acceptable response to seismic event: maintain operation, or safe shutdown only.", "C", "Design", "要確認"),
    ("Vendor's site supervisor shall be present on site during equipment installation and alignment.",
     "[CLARIFICATION] Site supervisor for installation and alignment: please confirm (a) duration of site assignment expected (Manufacturer estimates 3 weeks for alignment and pre-commissioning), (b) site supervisor cost basis (day rate, included in package price, or extra), (c) site safety induction requirements and lead time.", "C", "Cost / Delivery", "要確認"),
    ("All rotating equipment shall be dynamically balanced after each sub-assembly and after final assembly.",
     "[CLARIFICATION] Dynamic balancing at sub-assembly and final assembly is Manufacturer's standard practice for rotating equipment. Please confirm: (a) is a balance report to be included in the QC dossier (Manufacturer's standard: yes), (b) is the balance grade to be verified at final assembly speed or at a slower representative speed?", "C", "Quality", "要確認"),
    ("All instrument connections shall have isolation valves (block valves) and drain valves.",
     "[DEVIATION] Isolation and drain valves for all instruments significantly increase valve count and cost (+80-120 instrument root valves per package). Manufacturer proposes: block valves for all process transmitters (pressure, flow, temperature); no drain valves for clean utility services (cooling water, N2). Request approval of this approach.", "D", "Cost / Design", "対象"),
    ("Cooling water connections shall be titanium-lined for seawater cooling service.",
     "[DEVIATION] Titanium lining for cooling water connections is specified for seawater service. However, the design basis states the coolant is fresh water from a closed-loop cooling tower system. Titanium lining is unnecessary for fresh water and adds significant cost. If the coolant is confirmed as fresh water, Manufacturer proposes 316L SS as the tube material.", "D", "Cost / Design", "対象"),
    ("Vendor shall include a vibration isolation system (anti-vibration mounts) under the base frame.",
     "[CLARIFICATION] Vibration isolation mounts between base frame and foundation: please confirm (a) is a full dynamic analysis (FEA) of the mount system required, or is the mount selection based on static load plus API 610/617 allowance, (b) mount supplier preference (Mason Industries, Kinetics, or equivalent), (c) are seismic restraint clips required in addition to isolation mounts?", "C", "Design / Cost", "要確認"),
    ("Vendor shall submit a detailed energy consumption (power and utilities) summary.",
     "[CLARIFICATION] Utility consumption summary is provided in the bid technical data sheet. Please confirm (a) whether the summary should include part-load consumption data at 50%, 75%, and 100% flow, (b) the format required (tabular summary, or Sankey diagram energy balance), (c) whether cooling water flow at minimum and maximum ambient is required.", "C", "Design", "要確認"),
    ("All valves shall be manually operated unless specifically identified as automated on the P&ID.",
     "[CLARIFICATION] Manual valves as default, automated where specified: understood. Please finalize the P&ID with automation notations before Manufacturer completes the valve list, as the automation scope significantly affects the instrumentation BOM and cost. Target date for P&ID approval: please advise.", "C", "Design", "要確認"),
    ("Equipment nameplates shall be stainless steel (316SS) and attached by stainless steel screws.",
     "[CLARIFICATION] 316SS nameplates with SS screws are Manufacturer's standard. Please confirm: (a) nameplate content required (Manufacturer's standard: item tag, rated flow, rated head, speed, serial number, year of manufacture, Purchaser's PO number), (b) font size minimum (Manufacturer uses 3 mm minimum), (c) is a separate CE marking nameplate required?", "C", "Quality", "要確認"),
    ("The package shall be designed for hot-standby operation (standby unit fully pressurized).",
     "[CLARIFICATION] Hot-standby design is understood. Please confirm: (a) is the standby unit kept at full operating speed or at minimum governor speed, (b) automatic switchover upon trip of the operating unit — is this in scope (would require additional logic in the LCP and confirmation of parallel operation capability), (c) back-pressure check valve in Manufacturer's scope?", "C", "Design / Cost", "要確認"),
]

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = ["no", "要求仕様", "devicra_text", "type", "category", "デビクラ"]
COL_WIDTHS = [6, 60, 80, 8, 22, 12]

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_D      = PatternFill("solid", fgColor="FFF2CC")
FILL_C      = PatternFill("solid", fgColor="E2EFDA")
FILL_TGT    = PatternFill("solid", fgColor="FFDCE1")
FILL_CHK    = PatternFill("solid", fgColor="FFF9C4")
FILL_OUT    = PatternFill("solid", fgColor="E8F5E9")

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "past_devicra"

# Header row
for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = FILL_HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.row_dimensions[1].height = 22

# Data rows
for ri, (spec, dev, typ, cat, jdg) in enumerate(ROWS, start=2):
    vals = [ri - 1, spec, dev, typ, cat, jdg]
    for ci, v in enumerate(vals, start=1):
        cell = ws.cell(row=ri, column=ci, value=v)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = BORDER
        cell.font = Font(size=9)

    # row fill by type
    row_fill = FILL_D if typ == "D" else FILL_C
    ws.cell(row=ri, column=1).fill = row_fill
    ws.cell(row=ri, column=4).fill = row_fill

    # judgment fill
    jdg_fill = {"対象": FILL_TGT, "要確認": FILL_CHK, "対象外": FILL_OUT}.get(jdg)
    if jdg_fill:
        ws.cell(row=ri, column=6).fill = jdg_fill

    ws.row_dimensions[ri].height = 55

# Freeze panes and auto-filter
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS)+1}"

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "past_devicra_learning.xlsx")
wb.save(out)
print(f"Generated: {out}  ({len(ROWS)} rows)")
